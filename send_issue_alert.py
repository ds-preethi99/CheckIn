import requests
from datetime import datetime
from openpyxl import load_workbook
import calendar
from configparser import ConfigParser
import json

config = ConfigParser()                                 # Read the config file
config.read('config.ini')
discord_section = config['Discord']                     # Read the Discord section from the config file
endpoint = discord_section['ENDPOINT']                  # Get the webhook URL and user ID from the config file
token = discord_section['TOKEN']


def send_discord_message_exception(webhook_url, message_content):
    data = {
        "content": message_content
    }
    response = requests.post(webhook_url, data=data)
    return response


def send_individual_message_in_whatsapp(name_list, typename, logger):
    for name in name_list:
        # API endpoint and token
        logger.info('send whatsapp message to individual method executed')
        current_month = datetime.today().month  # get current month using the today().month function
        monthname = calendar.month_name[current_month]  # get current month name
        file_path = f'C:/Users/dsautomationslive/Employee Contact/{monthname}/Employee Contact.xlsx'
        employee_contact_book = load_workbook(file_path)
        employee_contact_sheet = employee_contact_book.active
        logger.info(f'get excel sheet {file_path}')
        phone_number = ''
        for cell in employee_contact_sheet[employee_contact_sheet.cell(row=1, column=18).column_letter]:
            if employee_contact_sheet.cell(row=cell.row, column=2).value == name:
                logger.info('get manager number')
                phone_number = employee_contact_sheet.cell(row=cell.row, column=3).value
                break
        logger.info(f'manager number {phone_number}')
        # Message data
        data = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": f"+91 {phone_number}",  # replace f"+91 {phone_number}"
            "type": "template",
            "template": {
                "name": "check_in_alert",
                "language": {
                    "code": "en_US"
                },
                "components": [
                    {
                        "type": "body",
                        "parameters": [
                            {"type": "text", "text": f"{name}"},
                            {"type": "text", "text": f"{typename}"}
                        ]
                    }
                ]
            }
        }
        headers = {
            'Authorization': f"Bearer {token}",
            'Content-Type': 'application/json'
        }
        response = requests.post(endpoint, headers=headers, data=json.dumps(data))
        if response.status_code == 200:
            logger.info(f'whatsapp message sent successfully')
        else:
            logger.error(f'whatsapp message didn\'t send for {phone_number}\n {response.text}')
