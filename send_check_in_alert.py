import os
import shutil
import send_issue_alert
from datetime import datetime
from openpyxl import load_workbook
from configparser import ConfigParser
import pandas as pd
config = ConfigParser()                                 # Read the config file
config.read('config.ini')
discord_section = config['Discord']                     # Read the Discord section from the config file
webhook_health_checkup_url = discord_section['WEBHOOK_HEALTH_CHECKUP_URL']


def move_attendance_file(logger, webhook_url):
    if not (os.path.exists('Attendance Tracker')):
        os.makedirs('Attendance Tracker')
    downloaded_file = 'C:/Users/dsautomationslive/Downloads/Attendance.xlsx'
    if os.path.exists(downloaded_file):
        shutil.move(downloaded_file, rf'Attendance Tracker/Attendance.xlsx')
        logger.info('Attendance file has been moved to Attendance Tracker successfully')
    else:
        logger.info('Attendance file doesn\'t exist')
        send_issue_alert.send_discord_message_exception(webhook_url,
                                                        'In check-in-check-out alert the Attendance file '
                                                        'haven\'t downloaded kindly check the logs')
        send_issue_alert.send_discord_message_exception(webhook_health_checkup_url,
                                                        'In check-in-check-out alert the Attendance file '
                                                        'haven\'t downloaded kindly check the logs')


def move_leave_register_file(logger, webhook_url):
    if not os.path.exists('Leave Register'):
        os.makedirs('Leave Register')
    downloaded_file = 'C:/Users/dsautomationslive/Downloads/Leave Register.xlsx'
    if os.path.exists(downloaded_file):
        shutil.move(downloaded_file, rf'Leave Register/Leave Register.xlsx')
        logger.info('Leave Register file has been moved to Leave Register folder successfully')
    else:
        logger.info('Leave Register file doesn\'t exist')
        send_issue_alert.send_discord_message_exception(webhook_url,
                                                        'In check-in-check-out alert the Attendance file '
                                                        'haven\'t downloaded kindly check the logs')
        send_issue_alert.send_discord_message_exception(webhook_health_checkup_url,
                                                        'In check-in-check-out alert the Attendance file '
                                                        'haven\'t downloaded kindly check the logs')


def send_checkin_alert(emp_name_list, lead_webhook_url, logger):
    today = datetime.today()
    current_date = today.date()
    formatted_date = current_date.strftime("%d/%m/%Y")
    content_message = (f'Check In Alert for {formatted_date}\nThe below listed people haven\'t check-in. '
                       f'Kindly check with them.\n')
    working_hours_df = pd.read_excel('Employee Working Hours/Employee Hour Report.xlsx')
    for name in emp_name_list:
        manager = working_hours_df.loc[working_hours_df['Employee Name'] == name, "Manager"].iloc[0]
        discord_id = working_hours_df.loc[working_hours_df['Manager'] == manager, "Discord ID"].iloc[0]
        content_message += f'<@{discord_id}>, {name} has not checked in\n'
    logger.info(f"{content_message}")
    send_issue_alert.send_discord_message_exception(lead_webhook_url, content_message)
    send_issue_alert.send_individual_message_in_whatsapp(emp_name_list, 'Check-in', logger)
    send_issue_alert.send_discord_message_exception('https://discord.com/api/webhooks/1160903527850332213/9A4YLzqCrFT75wSCpmcdEVcaNcQf80B36dwlHF135-DAm-VSAEByTWAdxUUw_HPifail', content_message)


def get_not_check_in_list(logger, lead_webhook_url, discord_user_id):
    # Constants for file paths
    monthName = datetime.now().strftime("%B")
    EMPLOYEE_CONTACT_PATH = f'C:/Users/dsautomationslive/Employee Contact/{monthName}/Employee Details.xlsx'
    ATTENDANCE_TRACKER_PATH = 'Attendance Tracker/Attendance.xlsx'
    LEAVE_REGISTER_PATH = 'Leave Register/Leave Register.xlsx'
    # Creating lists to store employee details
    empIdSet = set()
    empNameList = []
    empPhoneNumberList = []

    employee_manager_file_Path = EMPLOYEE_CONTACT_PATH
    attendance_file_Path = ATTENDANCE_TRACKER_PATH
    leave_register_file_Path = LEAVE_REGISTER_PATH

    # Check if files exist before loading workbooks
    if os.path.exists(employee_manager_file_Path):
        employee_manager_book = load_workbook(employee_manager_file_Path)
        employee_manager_sheet = employee_manager_book.active
        logger.info('employee details excel file is active')
    else:
        logger.critical('Employee details file not found')
        raise FileNotFoundError(f"File not found: {employee_manager_file_Path}")

    if os.path.exists(attendance_file_Path):
        attendance_book = load_workbook(attendance_file_Path)
        attendance_sheet = attendance_book.active
        logger.info('attendance excel file is active')
    else:
        logger.critical('attendance file not found')
        raise FileNotFoundError(f"File not found: {attendance_file_Path}")

    if os.path.exists(leave_register_file_Path):
        leave_register_book = load_workbook(leave_register_file_Path)
        leave_register_sheet = leave_register_book.active
        logger.info('leave register excel file is active')
        leave_ids = [cell.value for cell in leave_register_sheet['A'][1:] if cell.value is not None and leave_register_sheet.cell(row=cell.row, column=18).value in ['Applied','Approved']]
    else:
        logger.critical('leave register file not found')
        leave_ids = []  # Set leave_ids to an empty list if the file doesn't exist

    attendance_ids = [cell.value for cell in attendance_sheet['A'][1:]]
    logger.info(f'Leave : {leave_ids}\nattendance : {attendance_ids}')
    for cell in employee_manager_sheet[employee_manager_sheet.cell(row=2, column=13).column_letter]:
        if cell.row > 5:
            empId = employee_manager_sheet.cell(row=cell.row, column=1).value
            employement_status = employee_manager_sheet.cell(row=cell.row, column=10).value
            if employement_status in ['Active']:
                if empId in leave_ids or empId in ['DS001','DS002', '08', 'DS0154', 'DS058','DS174']:
                    logger.info(f'Leave List: {empId}')
                else:
                    if str(empId) in attendance_ids:
                        logger.info(f'{empId}')
                    else:
                        empName = employee_manager_sheet.cell(row=cell.row, column=2).value
                        empIdSet.add(empId)
                        empNameList.append(empName)
    if not empIdSet:
        logger.info(f'There is no record found for check-in')
    else:
        for cell in employee_manager_sheet[employee_manager_sheet.cell(row=5, column=2).column_letter]:
            if cell.row >= 5:
                for Id in empIdSet:
                    if Id == employee_manager_sheet.cell(row=cell.row, column=1).value:
                        empPhoneNumberList.append(employee_manager_sheet.cell(row=cell.row, column=4).value)
        employee_manager_book.close()       # Close the workbooks
        attendance_book.close()
        leave_register_book.close()
        logger.info(f'\nEmp ID : {empIdSet}\nEmp Name : {empNameList}\nEmp Phone Number: {empPhoneNumberList}')
        send_checkin_alert(empNameList, lead_webhook_url, logger)